from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtSql import QSqlDatabase, QSqlQuery

from PyQt5.QtWidgets import QFileDialog, QMessageBox

import sys
import openpyxl
import os
import re


def display_error_message(content):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setInformativeText(content)
    msg.setWindowTitle("Error")
    msg.exec_()

class Excel_Macro:
    def __init__(self, src_wb, dst_wb):
        self.src_wb = src_wb#openpyxl.load_workbook(src_path)
        self.dst_wb = dst_wb#openpyxl.load_workbook(dst_path)

    def copy_paste_range(self, copy_ws, paste_ws, copyfrom, pasteto):
        if copyfrom == "" or pasteto == "":
            pass
        elif ":" in copyfrom:
            for copy_row, paste_row in zip(copy_ws[copyfrom], paste_ws[pasteto]):
                for copy_cell, paste_cell in zip(copy_row, paste_row):
                    paste_cell.value = copy_cell.value
        else:
            paste_ws[pasteto].value = copy_ws[copyfrom].value

    def copy_pasteRange(src_sheet, dst_sheet, startCell, endCell):
        for row in src_sheet[startCell:endCell]:
            for cell in row:
                dst_sheet[cell.coordinate].value = cell.value

    def copyRange(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected
    def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        return countRow
    def createData(src_sheet, src_startCol, src_startRow, src_endCol, src_endRow, dst_sheet, dst_startCol, dst_startRow, dst_endCol, dst_endRow):
        print("Processing...")
        selectedRange = copyRange(src_startCol, src_startRow, src_endCol, src_endRow,src_sheet) 
        pastingRange = pasteRange(dst_startCol, dst_startRow, dst_endCol, dst_endRow,dst_sheet,selectedRange) 
        self.dst_wb.save(self.dst_wb)
        print("Range copied and pasted!")

class Dialog(QtWidgets.QDialog):
    def __init__(self, db, macro_id, parent=None):
        super(Dialog, self).__init__(parent)
        uic.loadUi("edit_dialog.ui", self)
        self.macro_id = macro_id
        self.sel_sheet_id = -1
        self.sel_cell_id = -1
        self.sheet_id_list = []
        self.cell_id_list = []
        self.db = db
        self.sheet_column = ['copy_from', 'paste_to']
        self.cell_column = ['copy_from', 'paste_to']
        self.macro_sheetview.setHorizontalHeaderLabels(['Copy from', 'Paste to'])
        self.macro_cellview.setHorizontalHeaderLabels(['Copy from', 'Paste to'])
        header = self.macro_sheetview.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        # header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header = self.macro_cellview.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        # header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        self.display_macro_info()
        self.manipulate_sheet_table()
        self.manipulate_cell_table()
        self.macro_sheetview.itemChanged.connect(self.sheet_changed)
        self.macro_cellview.itemChanged.connect(self.cell_changed)

    def display_macro_info(self):
        query = QSqlQuery()
        query.exec_("select * from Macro where id=" + str(self.macro_id))
        while(query.next()):
            self.macro_title.setText(query.value(1))
            self.macro_description.setText(query.value(2))


    def manipulate_sheet_table(self):       
        self.sheet_id_list = []
        self.macro_sheetview.setRowCount(0)
        query = QSqlQuery()
        query.exec_("select count(*) from Sheet where macro_id=" + str(self.macro_id))
        rows = 0
        while (query.next()):
            rows = int(query.value(0))
        # self.macro_sheetview.setItem(0, 0, QtWidgets.QTableWidgetItem("str(id)"))
        query = QSqlQuery()
        query.exec_("select id, copy_from, paste_to from Sheet where macro_id=" + str(self.macro_id))

        self.macro_sheetview.setRowCount(rows)
        row = 0
        while (query.next()):
            # id = int(query.value(0))
            # self.macro_sheetview.setItem(row, 0, QtWidgets.QTableWidgetItem(str(id)))
            self.sheet_id_list.append(int(query.value(0)))
            copy_from = str(query.value(1))
            self.macro_sheetview.setItem(row, 0, QtWidgets.QTableWidgetItem(str(copy_from)))
            paste_to = str(query.value(2))
            self.macro_sheetview.setItem(row, 1, QtWidgets.QTableWidgetItem(str(paste_to)))
            row += 1
    
    def manipulate_cell_table(self):
        self.cell_id_list = []
        self.macro_cellview.setRowCount(0)
        query = QSqlQuery()
        query.exec_("select count(*) from Cell where sheet_id=" + str(self.sel_sheet_id))
        rows = 0
        while (query.next()):
            rows = int(query.value(0))
        # self.macro_sheetview.setItem(0, 0, QtWidgets.QTableWidgetItem("str(id)"))
        query = QSqlQuery()
        query.exec_("select id, copy_from, paste_to from Cell where sheet_id=" + str(self.sel_sheet_id))

        self.macro_cellview.setRowCount(rows)
        row = 0
        while (query.next()):
            # id = int(query.value(0))
            # self.macro_sheetview.setItem(row, 0, QtWidgets.QTableWidgetItem(str(id)))
            self.cell_id_list.append(int(query.value(0)))
            copy_from = str(query.value(1))
            self.macro_cellview.setItem(row, 0, QtWidgets.QTableWidgetItem(str(copy_from)))
            paste_to = str(query.value(2))
            self.macro_cellview.setItem(row, 1, QtWidgets.QTableWidgetItem(str(paste_to)))
            row += 1
    @QtCore.pyqtSlot()
    def select_sheet_item(self):
        row = self.macro_sheetview.currentRow()       
        self.sel_sheet_id = self.sheet_id_list[row]
        self.manipulate_cell_table()

    

    @QtCore.pyqtSlot()
    def select_cell_item(self):
        row = self.macro_cellview.currentRow()
        self.sel_cell_id = self.cell_id_list[row]

    @QtCore.pyqtSlot()
    def delete_sheet_item(self):
        query = QSqlQuery()
        sql = "DELETE FROM Cell where sheet_id =" + str(self.sel_sheet_id)
        query.exec_(sql)
        sql = "DELETE FROM Sheet WHERE id =" + str(self.sel_sheet_id)
        query.exec_(sql)
        self.manipulate_sheet_table()
        self.manipulate_cell_table()

    @QtCore.pyqtSlot()
    def add_sheet_item(self):
        query = QSqlQuery()
        query.exec_("insert into Sheet (macro_id) VALUES('" + str(self.macro_id) + "')")
        self.manipulate_sheet_table()
    @QtCore.pyqtSlot()
    def add_cell_item(self):
        if self.sel_sheet_id == -1:
            QMessageBox.warning(self, "Warning", "Please select sheet entry. If there isn't any sheet entry, you have to put new sheet entry")
            return
        query = QSqlQuery()
        query.exec_("insert into Cell (sheet_id) VALUES('" + str(self.sel_sheet_id) + "')")
        self.manipulate_cell_table()
    
    @QtCore.pyqtSlot()
    def delete_cell_item(self):
        query = QSqlQuery()
        sql = "DELETE FROM Cell WHERE id =" + str(self.sel_cell_id)
        query.exec_(sql)
        self.manipulate_cell_table()

    @QtCore.pyqtSlot()
    def save_macro(self):
        query = QSqlQuery()
        sql = "update Macro set name = '" + str(self.macro_title.text()) +"', description = '" + str(self.macro_description.toPlainText()) + "' where id = " + str(self.macro_id)
        query.exec_(sql)
        self.accept()

    def sheet_changed(self, item):
        query = QSqlQuery()
        self.sel_sheet_id = self.sheet_id_list[int(item.row())]
        sql = "update Sheet set " + str(self.sheet_column[int(item.column())]) + " = '" + str(item.text())+"' where macro_id = " + str(self.macro_id) + " and id = " + str(self.sel_sheet_id)
        query.exec_(sql)

    def cell_changed(self, item):
        query = QSqlQuery()
        self.sel_cell_id = self.cell_id_list[int(item.row())]
        item_text = item.text()
        valid = True
        if item_text != "":
            for cell in item_text.split(":"):
                if not re.match("^[a-z^A-Z]+\d+$", cell):
                    display_error_message(cell + " isn't valid cell number")
                    return
        sql = "update Cell set " + str(self.cell_column[int(item.column())]) + " = '" + str(item.text())+"' where sheet_id = " + str(self.sel_sheet_id) + " and id = " + str(self.sel_cell_id)
        query.exec_(sql)

# macro_id_list = []
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        uic.loadUi("main.ui", self)
        # col_header = ["", ""]

        self.db = QSqlDatabase.addDatabase("QSQLITE")
        self.db.setDatabaseName("macro.db")
        self.db.open()
        create_table_Macro = """
            CREATE TABLE If NOT EXISTS "Macro" (
                "id"	INTEGER,
                "name"	TEXT,
                "description"	TEXT,
                PRIMARY KEY("id")
            )
        """
        create_table_Sheet = """
            CREATE TABLE If NOT EXISTS "Sheet"  (
                "id"	INTEGER PRIMARY KEY,
                "macro_id"	INTEGER,
                "copy_from"	TEXT,
                "paste_to"	TEXT,
                FOREIGN KEY (macro_id) REFERENCES Macro(id)
            )
        """
        create_table_Cell = """
            CREATE TABLE If NOT EXISTS "Cell" (
                'id'	INTEGER PRIMARY KEY,
                'sheet_id'	INTEGER,
                'copy_from'	TEXT,
                'paste_to'	TEXT,
                FOREIGN KEY (sheet_id) REFERENCES Sheet(id)
            )
        """
        query = QSqlQuery()
        query.exec_(create_table_Macro)
        query.exec_(create_table_Sheet)
        query.exec_(create_table_Cell)
        # self.db.close()
        # self.db.open()
        self.macro_id_list = []

        self.manipulate_macro_list()
        self.src_filename = ""
        self.dst_filename = ""
        self.sel_macro_id = -1

    def closeEvent(self, event):
        self.db.close()

    def manipulate_macro_list(self):
        self.list_macro.clear()
        # global macro_id_list
        self.macro_id_list = []
        query = QSqlQuery()
        query.exec_("select * from Macro")
        while (query.next()):
            id = int(query.value(0))
            self.macro_id_list.append(id)
            name = str(query.value(1))
            self.list_macro.addItem(name)
        
        # self.macro_id_list = macro_id_list
        
    @QtCore.pyqtSlot()
    def browser_src(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Excel Files (*.xlsx)", options=options)
        if fileName:
            self.src_filename = fileName
            self.edit_src.setText(fileName)

    @QtCore.pyqtSlot()
    def browser_dst(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","Excel Files (*.xlsx)", options=options)
        if fileName:
            self.dst_filename = fileName
            self.edit_dst.setText(fileName)

    @QtCore.pyqtSlot()
    def create_macro(self):
        query = QSqlQuery()
        sql = "INSERT into Macro (name, description) VALUES('new macro', 'new macro')"
        query.exec_(sql)
        self.manipulate_macro_list()
    @QtCore.pyqtSlot()
    def edit_macro(self):
        if self.sel_macro_id != -1:
            dialog = Dialog(self.db, self.sel_macro_id)
            if dialog.exec_():
                self.manipulate_macro_list()
            

    @QtCore.pyqtSlot()
    def remove_macro(self):        
        if self.sel_macro_id != -1:
            query = QSqlQuery()
            query.exec_("DELETE FROM Cell where sheet_id in (SELECT id from Sheet where Sheet.macro_id = " + str(self.sel_macro_id))
            query.exec_("DELETE from Sheet where macro_id = " + str(self.sel_macro_id))
            query.exec_("DELETE from Macro where id = " + str(self.sel_macro_id))
            self.manipulate_macro_list()
    @QtCore.pyqtSlot()
    def run_macro(self):
        if self.sel_macro_id != -1:
            if self.src_filename == "":
                display_error_message('Please Select Source file')
                return
            if self.dst_filename == "":
                display_error_message('Please Select Destination file')
                return
            
            src_wb = openpyxl.load_workbook(self.src_filename, data_only=True)
            dst_wb = openpyxl.load_workbook(self.dst_filename)
            if not os.path.exists(self.dst_filename):
                dst_wb = openpyxl.Workbook()
                ss_sheet = dst_wb['Sheet']
                ss_sheet.title = 'Sheet1'
                dst_wb.save(self.dst_filename)

            query = QSqlQuery()
            query.exec_("select * from Sheet where macro_id = " + str(self.sel_macro_id))
            while (query.next()):
                sheet_id = str(query.value(0))
                copy_ws = str(query.value(2))
                paste_ws = str(query.value(3))
                if not copy_ws in src_wb.sheetnames:
                    display_error_message("Don't exit the worksheet named " + copy_ws + "' in source file")
                    return
                if not paste_ws in dst_wb.sheetnames:
                    QMessageBox.information(self, "Information", "There isn't a worksheet named '" + paste_ws + "' in destination file, so will be created new sheet")
                    dst_wb.create_sheet(paste_ws)
                    dst_wb.save(self.dst_filename)
                query1 = QSqlQuery()
                query1.exec_("select * from Cell where sheet_id = " + sheet_id)
                while (query1.next()):
                    macro = Excel_Macro(src_wb, dst_wb)
                    copy_cell = str(query1.value(2))
                    paste_cell = str(query1.value(3))
                    macro.copy_paste_range(src_wb[copy_ws], dst_wb[paste_ws], copy_cell, paste_cell)
            dst_wb.save(self.dst_filename)
            QMessageBox.information(self, "Information", "Macro operation has been sucessed.")

        else:
            display_error_message('Please Selecte Macro')

    @QtCore.pyqtSlot()
    def select_macro(self):
        self.sel_macro_id = self.macro_id_list[int(self.list_macro.currentRow())]


if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication([])
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())
